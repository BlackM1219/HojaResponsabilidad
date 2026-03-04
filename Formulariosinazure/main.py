# -*- coding: utf-8 -*-
"""
Formulario DISAGRO - Sistema completo con integración a Teams
Versión con Power Automate Webhook

INSTRUCCIONES:
1. Instala dependencias: pip install openpyxl pywin32 requests
2. Configura tu WEBHOOK_URL (línea 30)
3. Asegúrate de tener el archivo plantilla.xlsx
4. ¡Ejecuta y usa!
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import win32com.client
from typing import Optional, Dict
from dataclasses import dataclass, field
import re
from datetime import datetime
import os
import requests
import base64

TEMPLATE_PATH = "https://default93a2cd4474774ba69ca3e4e7aaf9c2.dd.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/0f4c9b26740b401c9681e227af854f4e/triggers/manual/paths/invoke?api-version=1"

# ============================================
# 🔧 CONFIGURA AQUÍ TU URL DEL WEBHOOK
# ============================================
WEBHOOK_URL = "PEGA_TU_URL_AQUI"
# Ejemplo:
# WEBHOOK_URL = "https://prod-25.westus.logic.azure.com:443/workflows/abc123..."
# ============================================

@dataclass
class UserData:
    """Estructura de datos del usuario"""
    # Datos del usuario desde Outlook
    alias: str = ""
    full_name: str = ""
    first_name: str = ""
    second_name: str = ""
    first_surname: str = ""
    second_surname: str = ""
    correo: str = ""
    puesto: str = ""
    empresa: str = ""
    depart: str = ""
    
    # Datos del formulario
    ticket: str = ""
    fecha: str = ""
    localidad: str = "GT"
    pais: str = "GT"
    
    # Hardware
    pc_type: str = ""
    marca: str = ""
    model: str = ""
    st: str = ""
    fecha_compra: str = ""
    disco_duro: str = ""
    memoria_ram: str = ""
    ip: str = ""
    hostname: str = ""
    estado: str = "OK"
    
    # Equipos adicionales
    more_equipment: list = field(default_factory=list)
    temp_equipos: list = field(default_factory=list)
    
    # Software
    sistema_operativo: str = ""
    version_so: str = ""
    ofimática: str = ""
    version_office: str = ""
    
    # Problema y diagnóstico
    reported_problem: str = ""
    diagnosis: str = ""
    documentation: str = ""
    
    # Responsables
    tecnico: str = "Angel Lopez"
    usuario_firma: str = ""
    jefe_responsable: str = "Eder Morales"
    observaciones: str = ""
    
    # Tipos de formulario
    form_types: list = field(default_factory=list)

class TeamsWebhook:
    """Envía archivos a Teams usando Power Automate Webhook"""
    
    def __init__(self, webhook_url: str):
        self.webhook_url = webhook_url
    
    def enviar_archivo(self, file_path: str, email_destinatario: str, mensaje: str = "") -> bool:
        """Envía archivo a Teams mediante webhook de Power Automate"""
        
        # Verificar que el webhook esté configurado
        if "PEGA_TU_URL_AQUI" in self.webhook_url or not self.webhook_url.startswith("http"):
            messagebox.showerror(
                "⚠️ Webhook no configurado",
                "Debes configurar el WEBHOOK_URL en el código.\n\n"
                "Pasos:\n"
                "1. En Power Automate, abre tu Flow\n"
                "2. Haz clic en 'When a HTTP request is received'\n"
                "3. Copia la 'HTTP POST URL'\n"
                "4. Pégala en la variable WEBHOOK_URL (línea 30)\n\n"
                "La URL debe empezar con:\n"
                "https://prod-..."
            )
            return False
        
        try:
            # Leer el archivo y convertirlo a base64
            with open(file_path, 'rb') as f:
                file_content = f.read()
                file_base64 = base64.b64encode(file_content).decode('utf-8')
            
            file_name = os.path.basename(file_path)
            
            # Preparar datos para enviar (debe coincidir con el JSON Schema del Flow)
            payload = {
                "email": email_destinatario,
                "mensaje": mensaje or f"📋 Nuevo formulario: {file_name}",
                "archivo_nombre": file_name,
                "archivo_contenido": file_base64
            }
            
            print(f"📤 Enviando archivo a {email_destinatario}...")
            
            # Enviar al webhook
            response = requests.post(
                self.webhook_url,
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=30
            )
            
            if response.status_code in [200, 202, 204]:
                messagebox.showinfo(
                    "✅ Enviado exitosamente",
                    f"Archivo enviado a {email_destinatario}\n\n"
                    f"El destinatario recibirá el mensaje en Teams en unos segundos."
                )
                return True
            else:
                messagebox.showerror(
                    "❌ Error al enviar",
                    f"El servidor respondió con código {response.status_code}\n\n"
                    f"Detalles: {response.text[:200]}"
                )
                return False
                
        except FileNotFoundError:
            messagebox.showerror("❌ Error", f"No se encontró el archivo:\n{file_path}")
            return False
        except requests.exceptions.Timeout:
            messagebox.showerror("⏱️ Tiempo agotado", "El servidor tardó demasiado en responder.\nIntenta de nuevo.")
            return False
        except requests.exceptions.ConnectionError:
            messagebox.showerror("🌐 Error de conexión", "No se pudo conectar al servidor.\nVerifica tu conexión a internet.")
            return False
        except Exception as e:
            messagebox.showerror("❌ Error", f"Error al enviar:\n{str(e)}")
            return False

class OutlookSearcher:
    """Maneja la búsqueda en Outlook/GAL"""
    
    @staticmethod
    def buscar_usuario(alias: str) -> Optional[Dict[str, str]]:
        """Busca usuario en la Lista Global de Direcciones de Outlook"""
        alias = alias.lower().split("@")[0].strip()
        
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            ns = outlook.GetNamespace("MAPI")
            gal = ns.AddressLists["Lista global de direcciones"]
            
            for i in range(gal.AddressEntries.Count):
                entry = gal.AddressEntries.Item(i + 1)
                try:
                    ex = entry.GetExchangeUser()
                    correo = ex.PrimarySmtpAddress.lower()
                    
                    if correo.startswith(alias + "@"):
                        name_parts = ex.Name.split()
                        
                        first_name = ""
                        second_name = ""
                        first_surname = ""
                        second_surname = ""
                        
                        if len(name_parts) >= 1:
                            first_name = name_parts[0]
                        if len(name_parts) >= 3:
                            first_surname = name_parts[-2]
                            second_surname = name_parts[-1]
                            if len(name_parts) > 3:
                                second_name = " ".join(name_parts[1:-2])
                        elif len(name_parts) == 2:
                            first_surname = name_parts[1]
                        
                        return {
                            "alias": alias,
                            "full_name": ex.Name,
                            "first_name": first_name,
                            "second_name": second_name,
                            "first_surname": first_surname,
                            "second_surname": second_surname,
                            "correo": correo,
                            "puesto": ex.JobTitle or "",
                            "empresa": ex.CompanyName or "",
                            "depart": ex.Department or ""
                        }
                except AttributeError:
                    continue
                    
            return None
            
        except Exception as e:
            messagebox.showerror("Error COM", f"Error al conectar con Outlook:\n{str(e)}")
            return None

class ExcelGenerator:
    """Genera la plantilla de Excel con los datos"""
    
    FORM_TYPES = [
        "FECHA DE DEVOLUCIÓN DE EQUIPO",
        "HOJA DE RESPONSABILIDAD",
        "SOPORTE TÉCNICO",
        "PASE DE SALIDA",
        "DICTAMEN TÉCNICO",
        "RECEPCIÓN DE EQUIPO",
        "PRÉSTAMO DE EQUIPO"
    ]
    
    def __init__(self, template_path: str = TEMPLATE_PATH):
        self.template_path = template_path
    
    def generar(self, data: UserData, output_dir: str = ".") -> Optional[str]:
        """Genera la plantilla Excel con los datos proporcionados"""
        
        if not data.ticket or not data.alias:
            messagebox.showwarning(
                "Faltan datos",
                "Debes buscar un usuario y completar al menos el número de ticket."
            )
            return None
        
        try:
            wb = load_workbook(self.template_path)
            ws = wb.active
            
            if not data.fecha:
                data.fecha = datetime.now().strftime("%d/%m/%Y")
            
            replacements = self._preparar_datos(data)
            self._reemplazar_etiquetas(ws, replacements)
            self._marcar_checkboxes(ws, data)
            
            equipos = data.temp_equipos if data.temp_equipos else data.more_equipment
            if equipos:
                self._llenar_tabla_equipos(ws, equipos)
            
            safe_alias = re.sub(r'[^\w\-]', '_', data.alias)
            safe_ticket = re.sub(r'[^\w\-]', '_', data.ticket)
            output_file = os.path.join(output_dir, f"{safe_ticket}_{safe_alias}.xlsx")
            wb.save(output_file)
            
            messagebox.showinfo(
                "Guardado exitoso",
                f"Plantilla exportada como:\n{output_file}"
            )
            return output_file
            
        except FileNotFoundError:
            messagebox.showerror("Error", f"No se encontró la plantilla: {self.template_path}")
            return None
        except Exception as e:
            messagebox.showerror("Error al generar", f"Error al crear la plantilla:\n{str(e)}")
            return None
    
    def _preparar_datos(self, data: UserData) -> Dict[str, str]:
        """Prepara el diccionario de reemplazos"""
        return {
            "ticket": data.ticket,
            "fecha": data.fecha,
            "alias": data.alias,
            "first_name": data.first_name,
            "second_name": data.second_name,
            "first_surname": data.first_surname,
            "second_surname": data.second_surname,
            "puesto": data.puesto,
            "depart": data.depart,
            "empresa": data.empresa,
            "localidad": data.localidad,
            "pais": data.pais,
            "correo": data.correo,
            "pc_type": data.pc_type,
            "marca": data.marca,
            "model": data.model,
            "st": data.st,
            "fecha_compra": data.fecha_compra,
            "disco_duro": data.disco_duro,
            "memoria_ram": data.memoria_ram,
            "ip": data.ip,
            "hostname": data.hostname,
            "estado": data.estado,
            "sistema_operativo": data.sistema_operativo,
            "version_so": data.version_so,
            "ofimatica": data.ofimática,
            "version_office": data.version_office,
            "reported_problem": data.reported_problem,
            "diagnosis": data.diagnosis,
            "documentation": data.documentation,
            "tecnico": data.tecnico,
            "usuario_firma": data.usuario_firma or data.alias,
            "jefe_responsable": data.jefe_responsable,
            "observaciones": data.observaciones,
        }
    
    def _reemplazar_etiquetas(self, ws, replacements: Dict[str, str]):
        """Busca y reemplaza SOLO las etiquetas {{tag}} en la hoja"""
        pattern = re.compile(r'\{\{(\w+)\}\}')
        max_row = ws.max_row
        max_col = ws.max_column
        
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if cell.value and isinstance(cell.value, str) and "{{" in cell.value:
                        original_value = cell.value
                        new_value = original_value
                        
                        matches = pattern.findall(original_value)
                        
                        for tag in matches:
                            if tag in replacements:
                                placeholder = f"{{{{{tag}}}}}"
                                replacement_value = str(replacements[tag])
                                new_value = new_value.replace(placeholder, replacement_value)
                        
                        if new_value != original_value:
                            self._escribir_celda_segura(ws, row_idx, col_idx, new_value)
                
                except Exception:
                    continue
    
    def _escribir_celda_segura(self, ws, row: int, col: int, value):
        """Escribe en una celda manejando celdas combinadas"""
        try:
            celda = ws.cell(row=row, column=col)
            
            if isinstance(celda, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                        ws.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col,
                            value=value
                        )
                        return
            else:
                celda.value = value
        except Exception:
            try:
                ws.cell(row=row, column=col).value = value
            except:
                pass
    
    def _col_letter_to_index(self, letter: str) -> int:
        """Convierte letra de columna a índice numérico"""
        letter = letter.upper()
        idx = 0
        for ch in letter:
            idx = idx * 26 + (ord(ch) - ord('A') + 1)
        return idx
    
    def _marcar_checkboxes(self, ws, data: UserData):
        """Marca checkboxes de tipos de formulario y software"""
        
        form_checkboxes = {
            "HOJA DE RESPONSABILIDAD": ("B", 6),
            "SOPORTE TÉCNICO": ("M", 6),
            "SOPORTE TECNICO": ("M", 6),
            "PASE DE SALIDA": ("V", 6),
            "DICTAMEN TÉCNICO": ("B", 7),
            "DICTAMEN TECNICO": ("B", 7),
            "RECEPCIÓN DE EQUIPO": ("M", 7),
            "RECEPCION DE EQUIPO": ("M", 7),
            "PRÉSTAMO DE EQUIPO": ("V", 7),
            "PRESTAMO DE EQUIPO": ("V", 7),
        }
        
        if data.form_types:
            for form_type in data.form_types:
                if not form_type:
                    continue
                    
                key = form_type.strip().upper()
                key = key.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
                
                if key in form_checkboxes:
                    col_letter, row = form_checkboxes[key]
                    col_idx = self._col_letter_to_index(col_letter)
                    
                    try:
                        self._escribir_celda_segura(ws, row, col_idx, "X")
                    except Exception as e:
                        print(f"Error marcando checkbox {form_type}: {e}")
        
        if data.sistema_operativo:
            so = data.sistema_operativo.upper()
            if "WINDOWS" in so:
                col_idx = self._col_letter_to_index("J")
                self._escribir_celda_segura(ws, 38, col_idx, "X")
            if "MAC" in so:
                col_idx = self._col_letter_to_index("N")
                self._escribir_celda_segura(ws, 38, col_idx, "X")
        
        if data.ofimática:
            ofi = data.ofimática.upper()
            if "OFFICE" in ofi or "MS" in ofi:
                col_idx = self._col_letter_to_index("J")
                self._escribir_celda_segura(ws, 39, col_idx, "X")
    
    def _llenar_tabla_equipos(self, ws, equipos: list):
        """Llena la tabla de equipos adicionales en las celdas correctas"""
        if not equipos:
            return
        
        fila_inicio = 22
        
        col_equipo = self._col_letter_to_index("C")
        col_marca = self._col_letter_to_index("M")
        col_modelo = self._col_letter_to_index("U")
        col_serie = self._col_letter_to_index("AC")
        col_estado = self._col_letter_to_index("AO")
        
        for i, equipo_info in enumerate(equipos):
            row_num = fila_inicio + i
            try:
                self._escribir_celda_segura(ws, row_num, col_equipo, equipo_info.get("equipo", ""))
                self._escribir_celda_segura(ws, row_num, col_marca, equipo_info.get("marca", ""))
                self._escribir_celda_segura(ws, row_num, col_modelo, equipo_info.get("modelo", ""))
                self._escribir_celda_segura(ws, row_num, col_serie, equipo_info.get("serie", ""))
                self._escribir_celda_segura(ws, row_num, col_estado, equipo_info.get("estado", "OK"))
            except Exception as e:
                print(f"Error llenando equipo {i}: {e}")
                continue

class FormularioApp:
    """Aplicación principal con interfaz gráfica"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Formulario DISAGRO - Departamento de Sistemas")
        self.root.geometry("900x700")
        self.data = UserData()
        self.excel_gen = ExcelGenerator()
        self.teams_webhook = TeamsWebhook(WEBHOOK_URL)
        self.widgets = {}
        
        self._crear_interfaz()
    
    def _crear_interfaz(self):
        """Crea todos los elementos de la interfaz"""
        
        title_frame = tk.Frame(self.root, bg="#2c3e50", padx=10, pady=10)
        title_frame.grid(row=0, column=0, sticky="ew")
        
        tk.Label(
            title_frame,
            text="SISTEMA DE FORMULARIOS - DEPARTAMENTO DE SISTEMAS",
            font=("Arial", 14, "bold"),
            bg="#2c3e50",
            fg="white"
        ).pack()
        
        search_frame = tk.LabelFrame(self.root, text="1. Buscar Usuario", padx=10, pady=10)
        search_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        
        tk.Label(search_frame, text="Alias o correo:").grid(row=0, column=0, sticky="w")
        self.widgets['alias_entry'] = tk.Entry(search_frame, width=40)
        self.widgets['alias_entry'].grid(row=0, column=1, padx=5)
        
        # Permitir buscar con Enter
        self.widgets['alias_entry'].bind('<Return>', lambda e: self._buscar_usuario())
        
        tk.Button(
            search_frame,
            text="🔍 Buscar en Outlook",
            command=self._buscar_usuario,
            bg="#3498db",
            fg="white",
            padx=10
        ).grid(row=0, column=2)
        
        self._crear_treeview()
        
        button_frame = tk.Frame(self.root, padx=10, pady=10)
        button_frame.grid(row=3, column=0)
        
        # Crear botón de rellenar
        self.widgets['btn_rellenar'] = tk.Button(
            button_frame,
            text="📝 Rellenar Datos del Formulario",
            command=self._abrir_formulario_datos,
            state="disabled",  # Inicia deshabilitado
            bg="#e67e22",
            fg="white",
            padx=20,
            pady=10,
            font=("Arial", 10, "bold")
        )
        self.widgets['btn_rellenar'].grid(row=0, column=0, padx=5)
        
        tk.Button(
            button_frame,
            text="✅ Generar Plantilla Excel",
            command=self._generar_plantilla,
            bg="#27ae60",
            fg="white",
            padx=20,
            pady=10,
            font=("Arial", 10, "bold")
        ).grid(row=0, column=1, padx=5)
        
        tk.Button(
            button_frame,
            text="📤 Enviar a Teams",
            command=self._enviar_a_teams,
            bg="#0078d4",
            fg="white",
            padx=20,
            pady=10,
            font=("Arial", 10, "bold")
        ).grid(row=0, column=2, padx=5)
        
        # Debug: Imprimir para verificar
        print("✓ Interfaz creada correctamente")
        print(f"✓ Botón 'Rellenar' estado inicial: {self.widgets['btn_rellenar']['state']}")
    
    def _crear_treeview(self):
        """Crea el TreeView para mostrar resultados"""
        tree_frame = tk.LabelFrame(self.root, text="2. Usuario Encontrado", padx=10, pady=5)
        tree_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=5)
        
        columns = ("Nombre", "Correo", "Puesto", "Empresa", "Departamento")
        self.widgets['tree'] = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            height=3
        )
        
        for col in columns:
            self.widgets['tree'].heading(col, text=col)
            self.widgets['tree'].column(col, width=160)
        
        self.widgets['tree'].pack(fill="both", expand=True)
    
    def _buscar_usuario(self):
        """Busca usuario en Outlook"""
        alias = self.widgets['alias_entry'].get().strip()
        
        if not alias:
            messagebox.showwarning("Campo vacío", "Ingresa un alias o correo.")
            return
        
        tree = self.widgets['tree']
        tree.delete(*tree.get_children())
        
        print(f"🔍 Buscando usuario: {alias}")
        resultado = OutlookSearcher.buscar_usuario(alias)
        
        if resultado:
            print(f"✓ Usuario encontrado: {resultado['full_name']}")
            
            for key, value in resultado.items():
                setattr(self.data, key, value)
            
            tree.insert("", "end", values=(
                resultado["full_name"],
                resultado["correo"],
                resultado["puesto"],
                resultado["empresa"],
                resultado["depart"]
            ))
            
            # IMPORTANTE: Habilitar el botón
            print("⚙️ Habilitando botón 'Rellenar Datos'...")
            self.widgets['btn_rellenar'].config(state="normal")
            print(f"✓ Estado del botón ahora: {self.widgets['btn_rellenar']['state']}")
            
            messagebox.showinfo("✓ Usuario encontrado", f"Usuario {resultado['full_name']} cargado.")
        else:
            print("❌ Usuario no encontrado")
            # Deshabilitar el botón si no se encuentra
            self.widgets['btn_rellenar'].config(state="disabled")
            messagebox.showwarning("No encontrado", f"No se encontró el usuario: {alias}")
    
    def _abrir_formulario_datos(self):
        """Abre ventana para capturar datos adicionales"""
        
        print("📝 Abriendo ventana de formulario...")
        
        if not self.data.alias:
            messagebox.showwarning("Busca primero", "Primero debes buscar un usuario.")
            return
        
        print(f"✓ Usuario cargado: {self.data.full_name}")
        
        win = tk.Toplevel(self.root)
        win.title(f"Datos del Formulario - {self.data.full_name}")
        win.geometry("800x900")
        
        canvas = tk.Canvas(win)
        scrollbar = tk.Scrollbar(win, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        entries = {}
        row = 0
        
        # ENCABEZADO
        tk.Label(scrollable_frame, text="DATOS DEL FORMULARIO", font=("Arial", 12, "bold"), 
                 bg="#2c3e50", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        row += 1
        
        # Ticket y fecha
        tk.Label(scrollable_frame, text="No. de Ticket *").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['ticket'] = tk.Entry(scrollable_frame, width=50)
        entries['ticket'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['ticket'].insert(0, self.data.ticket)
        row += 1
        
        tk.Label(scrollable_frame, text="Fecha").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['fecha'] = tk.Entry(scrollable_frame, width=50)
        entries['fecha'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['fecha'].insert(0, datetime.now().strftime("%d/%m/%Y"))
        row += 1
        
        # TIPO DE FORMULARIO
        tk.Label(scrollable_frame, text="TIPO DE FORMULARIO", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        tk.Label(scrollable_frame, text="Selecciona tipo(s) *").grid(row=row, column=0, sticky="ne", padx=5, pady=5)
        
        form_frame = tk.Frame(scrollable_frame)
        form_frame.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        
        form_vars = {}
        for i, form_type in enumerate(ExcelGenerator.FORM_TYPES):
            var = tk.BooleanVar()
            if form_type in self.data.form_types:
                var.set(True)
            cb = tk.Checkbutton(form_frame, text=form_type, variable=var)
            cb.grid(row=i, column=0, sticky="w")
            form_vars[form_type] = var
        
        entries['form_types'] = form_vars
        row += 1
        
        # SECCIÓN A: DATOS DEL COLABORADOR
        tk.Label(scrollable_frame, text="A. DATOS DEL COLABORADOR", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        tk.Label(scrollable_frame, text="(Ya cargados desde Outlook)", font=("Arial", 9, "italic"), 
                 fg="gray").grid(row=row, column=0, columnspan=2, pady=2)
        row += 1
        
        campos_a = [
            ("Localidad", "localidad"),
            ("País", "pais"),
        ]
        
        for label, key in campos_a:
            tk.Label(scrollable_frame, text=label).grid(row=row, column=0, sticky="e", padx=5, pady=5)
            entries[key] = tk.Entry(scrollable_frame, width=50)
            entries[key].grid(row=row, column=1, sticky="w", padx=5, pady=5)
            entries[key].insert(0, getattr(self.data, key, ""))
            row += 1
        
        # SECCIÓN B: HARDWARE
        tk.Label(scrollable_frame, text="B. HARDWARE", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        campos_hardware = [
            ("Tipo de Computadora *", "pc_type"),
            ("Marca *", "marca"),
            ("Modelo", "model"),
            ("Service Tag (ST) *", "st"),
            ("Fecha de Compra", "fecha_compra"),
            ("Disco Duro", "disco_duro"),
            ("Memoria RAM", "memoria_ram"),
            ("IP", "ip"),
            ("Host Name", "hostname"),
        ]
        
        for label, key in campos_hardware:
            tk.Label(scrollable_frame, text=label).grid(row=row, column=0, sticky="e", padx=5, pady=5)
            entries[key] = tk.Entry(scrollable_frame, width=50)
            entries[key].grid(row=row, column=1, sticky="w", padx=5, pady=5)
            entries[key].insert(0, getattr(self.data, key, ""))
            row += 1
        
        # Estado del equipo
        tk.Label(scrollable_frame, text="Estado del equipo").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['estado'] = ttk.Combobox(
            scrollable_frame,
            values=["OK", "MALO", "REGULAR"],
            state="readonly",
            width=47
        )
        entries['estado'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['estado'].set(self.data.estado)
        row += 1
        
        # SECCIÓN C: EQUIPOS ADICIONALES
        tk.Label(scrollable_frame, text="C. EQUIPOS ADICIONALES", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        equipos_frame = tk.Frame(scrollable_frame, relief="sunken", bd=1)
        equipos_frame.grid(row=row, column=0, columnspan=2, padx=5, pady=5, sticky="ew")
        
        equipos_list = tk.Listbox(equipos_frame, height=5, width=80)
        equipos_list.pack(side="left", fill="both", expand=True)
        
        equipos_scroll = tk.Scrollbar(equipos_frame, command=equipos_list.yview)
        equipos_scroll.pack(side="right", fill="y")
        equipos_list.config(yscrollcommand=equipos_scroll.set)
        
        if self.data.more_equipment:
            for eq in self.data.more_equipment:
                equipos_list.insert("end", f"{eq.get('equipo', '')} - {eq.get('marca', '')} {eq.get('modelo', '')} - {eq.get('serie', '')}")
        
        btn_equipos_frame = tk.Frame(scrollable_frame)
        btn_equipos_frame.grid(row=row+1, column=0, columnspan=2)
        
        def agregar_equipo():
            eq_win = tk.Toplevel(win)
            eq_win.title("Agregar Equipo Adicional")
            eq_win.geometry("450x300")
            
            eq_entries = {}
            labels = ["Equipo:", "Marca:", "Modelo:", "Serie:", "Estado:"]
            keys = ["equipo", "marca", "modelo", "serie", "estado"]
            
            for i, (lbl, key) in enumerate(zip(labels, keys)):
                tk.Label(eq_win, text=lbl).grid(row=i, column=0, sticky="e", padx=5, pady=5)
                if key == "estado":
                    e = ttk.Combobox(eq_win, values=["OK", "MALO", "REGULAR"], state="readonly", width=28)
                    e.set("OK")
                else:
                    e = tk.Entry(eq_win, width=30)
                e.grid(row=i, column=1, padx=5, pady=5)
                eq_entries[key] = e
            
            def guardar_equipo():
                equipo_data = {}
                for k, v in eq_entries.items():
                    equipo_data[k] = v.get().strip()
                
                if equipo_data["equipo"]:
                    equipos_list.insert("end", f"{equipo_data['equipo']} - {equipo_data['marca']} {equipo_data['modelo']} - {equipo_data['serie']}")
                    if not hasattr(self.data, 'temp_equipos'):
                        self.data.temp_equipos = []
                    self.data.temp_equipos.append(equipo_data)
                    eq_win.destroy()
                else:
                    messagebox.showwarning("Campo vacío", "El nombre del equipo es obligatorio")
            
            tk.Button(eq_win, text="💾 Guardar", command=guardar_equipo, bg="#27ae60", fg="white").grid(row=len(labels), column=0, columnspan=2, pady=10)
        
        def eliminar_equipo():
            sel = equipos_list.curselection()
            if not sel:
                messagebox.showwarning("Selecciona", "Selecciona un equipo para eliminar")
                return
            idx = sel[0]
            equipos_list.delete(idx)
            if hasattr(self.data, 'temp_equipos') and idx < len(self.data.temp_equipos):
                try:
                    self.data.temp_equipos.pop(idx)
                except:
                    pass
        
        tk.Button(btn_equipos_frame, text="➕ Agregar", command=agregar_equipo, bg="#3498db", fg="white").grid(row=0, column=0, padx=5, pady=5)
        tk.Button(btn_equipos_frame, text="➖ Eliminar", command=eliminar_equipo, bg="#c0392b", fg="white").grid(row=0, column=1, padx=5, pady=5)
        
        row += 3
        
        # SECCIÓN D: SOFTWARE
        tk.Label(scrollable_frame, text="D. SOFTWARE", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        # Sistema operativo con checkboxes
        tk.Label(scrollable_frame, text="Sistema operativo").grid(row=row, column=0, sticky="ne", padx=5, pady=5)
        so_frame = tk.Frame(scrollable_frame)
        so_frame.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        
        so_vars = {}
        so_options = ["Windows", "Mac OS"]
        for i, so_opt in enumerate(so_options):
            var = tk.BooleanVar()
            if so_opt in self.data.sistema_operativo:
                var.set(True)
            cb = tk.Checkbutton(so_frame, text=so_opt, variable=var)
            cb.grid(row=0, column=i, sticky="w", padx=10)
            so_vars[so_opt] = var
        
        entries['sistema_operativo_vars'] = so_vars
        row += 1
        
        # Versión de Sistema Operativo
        tk.Label(scrollable_frame, text="Versión SO").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['version_so'] = tk.Entry(scrollable_frame, width=50)
        entries['version_so'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['version_so'].insert(0, getattr(self.data, 'version_so', ''))
        row += 1
        
        # Ofimática con checkboxes
        tk.Label(scrollable_frame, text="Ofimática").grid(row=row, column=0, sticky="ne", padx=5, pady=5)
        ofi_frame = tk.Frame(scrollable_frame)
        ofi_frame.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        
        ofi_vars = {}
        ofi_options = ["MS Office", "Open Office"]
        for i, ofi_opt in enumerate(ofi_options):
            var = tk.BooleanVar()
            if ofi_opt in self.data.ofimática:
                var.set(True)
            cb = tk.Checkbutton(ofi_frame, text=ofi_opt, variable=var)
            cb.grid(row=0, column=i, sticky="w", padx=10)
            ofi_vars[ofi_opt] = var
        
        entries['ofimatica_vars'] = ofi_vars
        row += 1
        
        # Versión de Office
        tk.Label(scrollable_frame, text="Versión Office").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['version_office'] = tk.Entry(scrollable_frame, width=50)
        entries['version_office'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['version_office'].insert(0, getattr(self.data, 'version_office', ''))
        row += 1
        
        # SECCIÓN E: PROBLEMA REPORTADO
        tk.Label(scrollable_frame, text="E. PROBLEMA REPORTADO", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        tk.Label(scrollable_frame, text="Problema reportado:").grid(row=row, column=0, sticky="ne", padx=5, pady=5)
        txt_problem = tk.Text(scrollable_frame, width=50, height=4)
        txt_problem.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        if self.data.reported_problem:
            txt_problem.insert("1.0", self.data.reported_problem)
        entries['reported_problem'] = txt_problem
        row += 1
        
        # SECCIÓN F: DIAGNÓSTICO
        tk.Label(scrollable_frame, text="F. DIAGNÓSTICO", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        tk.Label(scrollable_frame, text="Diagnóstico:").grid(row=row, column=0, sticky="ne", padx=5, pady=5)
        txt_diag = tk.Text(scrollable_frame, width=50, height=6)
        txt_diag.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        if self.data.diagnosis:
            txt_diag.insert("1.0", self.data.diagnosis)
        entries['diagnosis'] = txt_diag
        row += 1
        
        tk.Label(scrollable_frame, text="Documentación/Repuestos:").grid(row=row, column=0, sticky="ne", padx=5, pady=5)
        txt_doc = tk.Text(scrollable_frame, width=50, height=3)
        txt_doc.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        if self.data.documentation:
            txt_doc.insert("1.0", self.data.documentation)
        entries['documentation'] = txt_doc
        row += 1
        
        # Botón de guardar
        def guardar_formulario():
            ticket_val = entries['ticket'].get().strip()
            if not ticket_val:
                messagebox.showwarning("Faltan datos", "El número de ticket es obligatorio")
                return
            
            self.data.ticket = ticket_val
            self.data.fecha = entries['fecha'].get().strip()
            
            selected_forms = []
            for k, var in entries['form_types'].items():
                if var.get():
                    selected_forms.append(k)
            self.data.form_types = selected_forms
            
            self.data.localidad = entries['localidad'].get().strip()
            self.data.pais = entries['pais'].get().strip()
            
            for label, key in campos_hardware:
                setattr(self.data, key, entries[key].get().strip())
            self.data.estado = entries['estado'].get().strip()
            
            selected_so = [so for so, var in entries['sistema_operativo_vars'].items() if var.get()]
            self.data.sistema_operativo = ", ".join(selected_so)
            self.data.version_so = entries['version_so'].get().strip()
            
            selected_ofi = [ofi for ofi, var in entries['ofimatica_vars'].items() if var.get()]
            self.data.ofimática = ", ".join(selected_ofi)
            self.data.version_office = entries['version_office'].get().strip()
            
            self.data.reported_problem = entries['reported_problem'].get("1.0", "end").strip()
            self.data.diagnosis = entries['diagnosis'].get("1.0", "end").strip()
            self.data.documentation = entries['documentation'].get("1.0", "end").strip()
            
            print("✓ Datos del formulario guardados")
            messagebox.showinfo("Guardado", "Datos del formulario guardados en memoria.")
            win.destroy()
        
        tk.Button(scrollable_frame, text="💾 Guardar formulario", command=guardar_formulario, 
                 bg="#27ae60", fg="white", padx=10, pady=8, font=("Arial", 10, "bold")).grid(row=row, column=0, columnspan=2, pady=15)
        
        print("✓ Ventana de formulario creada")
    
    def _generar_plantilla(self):
        """Genera la plantilla Excel"""
        if not self.data.alias:
            messagebox.showwarning("Busca primero", "Primero debes buscar un usuario.")
            return
        if not self.data.ticket:
            messagebox.showwarning("Falta ticket", "Debes completar el número de ticket.")
            return
        
        out_dir = filedialog.askdirectory(title="Selecciona carpeta para guardar", initialdir=os.getcwd())
        if not out_dir:
            if not messagebox.askyesno("Continuar", "No seleccionaste carpeta. ¿Deseas guardar en la carpeta actual?"):
                return
            out_dir = os.getcwd()
        
        output_file = self.excel_gen.generar(self.data, output_dir=out_dir)
        if output_file:
            self.ultimo_archivo = output_file
            self.data.temp_equipos = []
    
    def _enviar_a_teams(self):
        """Envía el archivo generado a un usuario de Teams usando Webhook"""
        if not hasattr(self, 'ultimo_archivo') or not os.path.exists(getattr(self, 'ultimo_archivo', '')):
            respuesta = messagebox.askyesno(
                "Archivo no generado",
                "No hay un archivo generado recientemente.\n¿Deseas generar uno ahora?"
            )
            if respuesta:
                self._generar_plantilla()
                if not hasattr(self, 'ultimo_archivo'):
                    return
            else:
                return
        
        teams_win = tk.Toplevel(self.root)
        teams_win.title("Enviar a Microsoft Teams")
        teams_win.geometry("500x350")
        teams_win.resizable(False, False)
        
        main_frame = tk.Frame(teams_win, padx=20, pady=20)
        main_frame.pack(fill="both", expand=True)
        
        tk.Label(
            main_frame,
            text="📤 Enviar a Teams (Power Automate)",
            font=("Arial", 14, "bold"),
            fg="#0078d4"
        ).pack(pady=(0, 20))
        
        tk.Label(main_frame, text="Archivo:", font=("Arial", 10, "bold")).pack(anchor="w")
        tk.Label(
            main_frame,
            text=os.path.basename(self.ultimo_archivo),
            font=("Arial", 9),
            fg="gray",
            wraplength=450
        ).pack(anchor="w", pady=(0, 15))
        
        tk.Label(main_frame, text="Email del destinatario:", font=("Arial", 10, "bold")).pack(anchor="w")
        email_entry = tk.Entry(main_frame, width=50, font=("Arial", 10))
        email_entry.pack(fill="x", pady=(5, 15))
        
        if self.data.correo:
            email_entry.insert(0, self.data.correo)
        
        tk.Label(main_frame, text="Mensaje (opcional):", font=("Arial", 10, "bold")).pack(anchor="w")
        mensaje_text = tk.Text(main_frame, width=50, height=3, font=("Arial", 9))
        mensaje_text.pack(fill="x", pady=(5, 15))
        
        mensaje_default = f"📋 Formulario de {self.data.full_name} - Ticket #{self.data.ticket}"
        mensaje_text.insert("1.0", mensaje_default)
        
        btn_frame = tk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=(10, 0))
        
        def enviar():
            email = email_entry.get().strip()
            if not email:
                messagebox.showwarning("Email requerido", "Debes ingresar el email del destinatario")
                return
            
            if "@" not in email:
                messagebox.showwarning("Email inválido", "Ingresa un email válido")
                return
            
            mensaje = mensaje_text.get("1.0", "end").strip()
            
            teams_win.config(cursor="wait")
            teams_win.update()
            
            try:
                exito = self.teams_webhook.enviar_archivo(
                    self.ultimo_archivo,
                    email,
                    mensaje
                )
                
                teams_win.config(cursor="")
                
                if exito:
                    teams_win.destroy()
            except Exception as e:
                teams_win.config(cursor="")
                messagebox.showerror("Error", f"Error al enviar:\n{str(e)}")
        
        tk.Button(
            btn_frame,
            text="📤 Enviar",
            command=enviar,
            bg="#0078d4",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=30,
            pady=8,
            cursor="hand2"
        ).pack(side="left", padx=(0, 10))
        
        tk.Button(
            btn_frame,
            text="❌ Cancelar",
            command=teams_win.destroy,
            bg="#666",
            fg="white",
            font=("Arial", 10),
            padx=30,
            pady=8,
            cursor="hand2"
        ).pack(side="left")

def main():
    root = tk.Tk()
    app = FormularioApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
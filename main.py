# -*- coding: utf-8 -*-
"""
Formulario DISAGRO - Código fusionado
- Mantiene todas las funciones originales que enviaste.
- Añade mejoras no destructivas:
  * Mapeo explícito para marcar celdas de tipos de formulario y software.
  * Guardado local configurable.
  * Stub para enviar archivo a Teams (documentado).
  * Conserva la lógica original de búsqueda en Outlook, reemplazos y tablas.
Marcas:
  # --- ORIGINAL ---  -> código que proviene de tu versión original
  # --- ADICIONAL / STUB --- -> nuevas funciones o extensiones añadidas
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import win32com.client
from typing import Optional, Dict
from dataclasses import dataclass, field
import re
from datetime import datetime
import os

TEMPLATE_PATH = "plantilla.xlsx"

# --- ORIGINAL: UserData dataclass (conservado) ---
@dataclass
class UserData:
    """Estructura de datos del usuario"""
    # Datos del usuario desde Outlook
    alias: str = ""
    full_name: str = ""
    first_name: str = ""
    second_name: str = ""
    first_surname: str = ""
    second_surname: str = ""
    correo: str = ""
    puesto: str = ""
    empresa: str = ""
    depart: str = ""
    
    # Sección A: Datos del colaborador (campos 1-10)
    localidad: str = "GT"  # Campo 9
    pais: str = "GT"  # Campo 10
    
    # Datos del formulario
    ticket: str = ""
    fecha: str = ""
    
    # Sección B: Hardware (campos 11-19)
    pc_type: str = ""  # Campo 11: Tipo de computadora
    marca: str = ""  # Campo 12: Marca
    model: str = ""  # Campo 13: Modelo
    st: str = ""  # Campo 14: Service Tag
    fecha_compra: str = ""  # Campo 15: Fecha de compra
    disco_duro: str = ""  # Campo 16
    memoria_ram: str = ""  # Campo 17
    ip: str = ""  # Campo 18
    hostname: str = ""  # Campo 19
    
    # Sección C: Equipos adicionales
    more_equipment: list = field(default_factory=list)
    
    # Sección D: Software
    sistema_operativo: str = ""  # Windows/Mac OS
    ofimática: str = ""  # MS Office/Open Office
    software_otros: list = field(default_factory=list)  # Lista de software adicional
    
    # Sección E, F: Problema y diagnóstico
    reported_problem: str = ""
    diagnosis: str = ""
    documentation: str = ""
    
    # Sección G: Responsables
    tecnico: str = "Angel Lopez"  # Por defecto
    usuario_firma: str = ""
    jefe_responsable: str = "Eder Morales"  # Por defecto
    observaciones: str = ""
    
    # Tipos de formulario seleccionados
    form_types: list = field(default_factory=list)
    estado: str = "OK"  # Estado del equipo

    # Temporal para equipos añadidos en UI (ADICIONAL)
    temp_equipos: list = field(default_factory=list)

# --- ORIGINAL: OutlookSearcher (conservado) ---
class OutlookSearcher:
    """Maneja la búsqueda en Outlook/GAL"""
    
    @staticmethod
    def buscar_usuario(alias: str) -> Optional[Dict[str, str]]:
        """Busca usuario en la Lista Global de Direcciones de Outlook"""
        alias = alias.lower().split("@")[0].strip()
        
        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            ns = outlook.GetNamespace("MAPI")
            gal = ns.AddressLists["Lista global de direcciones"]
            
            for i in range(gal.AddressEntries.Count):
                entry = gal.AddressEntries.Item(i + 1)
                try:
                    ex = entry.GetExchangeUser()
                    correo = ex.PrimarySmtpAddress.lower()
                    
                    if correo.startswith(alias + "@"):
                        name_parts = ex.Name.split()
                        
                        first_name = ""
                        second_name = ""
                        first_surname = ""
                        second_surname = ""
                        
                        if len(name_parts) >= 1:
                            first_name = name_parts[0]
                        if len(name_parts) >= 3:
                            first_surname = name_parts[-2]
                            second_surname = name_parts[-1]
                            if len(name_parts) > 3:
                                second_name = " ".join(name_parts[1:-2])
                        elif len(name_parts) == 2:
                            first_surname = name_parts[1]
                        
                        return {
                            "alias": alias,
                            "full_name": ex.Name,
                            "first_name": first_name,
                            "second_name": second_name,
                            "first_surname": first_surname,
                            "second_surname": second_surname,
                            "correo": correo,
                            "puesto": ex.JobTitle or "",
                            "empresa": ex.CompanyName or "",
                            "depart": ex.Department or ""
                        }
                except AttributeError:
                    continue
                    
            return None
            
        except Exception as e:
            messagebox.showerror("Error COM", f"Error al conectar con Outlook:\n{str(e)}")
            return None

# --- ORIGINAL: ExcelGenerator base (conservado y extendido) ---
class ExcelGenerator:
    """Genera la plantilla de Excel con los datos"""
    
    FORM_TYPES = [
        "FECHA DE DEVOLUCIÓN DE EQUIPO",
        "HOJA DE RESPONSABILIDAD",
        "SOPORTE TÉCNICO",
        "PASE DE SALIDA",
        "DICTAMEN TÉCNICO",
        "RECEPCIÓN DE EQUIPO",
        "PRÉSTAMO DE EQUIPO"
    ]
    
    def __init__(self, template_path: str = TEMPLATE_PATH):
        self.template_path = template_path
    
    # --- ORIGINAL: generar (ligeramente adaptado para devolver ruta) ---
    def generar(self, data: UserData, output_dir: str = ".") -> Optional[str]:
        """Genera la plantilla Excel con los datos proporcionados y devuelve la ruta del archivo"""
        
        if not data.ticket or not data.alias:
            messagebox.showwarning(
                "Faltan datos",
                "Debes buscar un usuario y completar al menos el número de ticket."
            )
            return None
        
        try:
            wb = load_workbook(self.template_path)
            ws = wb.active
            
            # Agregar fecha automática
            if not data.fecha:
                data.fecha = datetime.now().strftime("%d/%m/%Y")
            
            # Preparar datos y reemplazar etiquetas
            replacements = self._preparar_datos(data)
            self._reemplazar_etiquetas(ws, replacements)
            
            # --- ADICIONAL: marcar formularios y software en celdas específicas (no destructivo) ---
            self._marcar_formularios_y_software(ws, data)
            
            # Llenar tabla de equipos adicionales (usar temp_equipos si existen)
            equipos = data.temp_equipos if getattr(data, "temp_equipos", None) else data.more_equipment
            if equipos:
                self._llenar_tabla_equipos(ws, equipos)
            
            # Llenar tabla de software
            if data.software_otros:
                self._llenar_tabla_software(ws, data.software_otros)
            
            # Guardar archivo (nombre seguro)
            safe_alias = re.sub(r'[^\w\-]', '_', data.alias)
            safe_ticket = re.sub(r'[^\w\-]', '_', data.ticket)
            output_file = os.path.join(output_dir, f"{safe_ticket}_{safe_alias}.xlsx")
            wb.save(output_file)
            
            return output_file
            
        except FileNotFoundError:
            messagebox.showerror(
                "Error",
                f"No se encontró la plantilla: {self.template_path}"
            )
            return None
        except Exception as e:
            messagebox.showerror(
                "Error al generar",
                f"Error al crear la plantilla:\n{str(e)}"
            )
            return None
    
    # --- ORIGINAL: _preparar_datos (conservado) ---
    def _preparar_datos(self, data: UserData) -> Dict[str, str]:
        """Prepara el diccionario de reemplazos según tu plantilla"""
        return {
            # Encabezado
            "ticket": data.ticket,
            "fecha": data.fecha,
            
            # Sección A: Datos del colaborador (campos 1-10)
            "alias": data.alias,
            "first_name": data.first_name,
            "second_name": data.second_name,
            "first_surname": data.first_surname,
            "second_surname": data.second_surname,
            "puesto": data.puesto,
            "depart": data.depart,
            "empresa": data.empresa,
            "localidad": data.localidad,
            "pais": data.pais,
            
            # Sección B: Hardware (campos 11-19)
            "pc_type": data.pc_type,
            "marca": data.marca,
            "model": data.model,
            "st": data.st,
            "fecha_compra": data.fecha_compra,
            "disco_duro": data.disco_duro,
            "memoria_ram": data.memoria_ram,
            "ip": data.ip,
            "hostname": data.hostname,
            
            # Sección D: Software
            "sistema_operativo": data.sistema_operativo,
            "ofimatica": data.ofimática,
            
            # Sección E, F
            "reported_problem": data.reported_problem,
            "diagnosis": data.diagnosis,
            "documentation": data.documentation,
            
            # Sección G: Responsables
            "tecnico": data.tecnico,
            "usuario_firma": data.usuario_firma or data.alias,
            "jefe_responsable": data.jefe_responsable,
            "observaciones": data.observaciones,
            
            # Estado
            "estado": data.estado,
        }
    
    # --- ORIGINAL: _reemplazar_etiquetas (conservado) ---
    def _reemplazar_etiquetas(self, ws, replacements: Dict[str, str]):
        """Busca y reemplaza todas las etiquetas {{tag}} en la hoja"""
        pattern = re.compile(r'\{\{(\w+)\}\}')
        max_row = ws.max_row
        max_col = ws.max_column
        
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if cell.value and isinstance(cell.value, str):
                        original_value = cell.value
                        new_value = original_value
                        
                        matches = pattern.findall(original_value)
                        
                        for tag in matches:
                            if tag in replacements:
                                placeholder = f"{{{{{tag}}}}}"
                                replacement_value = str(replacements[tag])
                                new_value = new_value.replace(placeholder, replacement_value)
                        
                        if new_value != original_value:
                            self._escribir_celda_segura(ws, row_idx, col_idx, new_value)
                
                except Exception:
                    continue
    
    # --- ORIGINAL: _escribir_celda_segura (conservado) ---
    def _escribir_celda_segura(self, ws, row: int, col: int, value: str):
        """Escribe en una celda manejando celdas combinadas"""
        try:
            celda = ws.cell(row=row, column=col)
            
            if isinstance(celda, MergedCell):
                for merged_range in ws.merged_cells.ranges:
                    if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                        ws.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col,
                            value=value
                        )
                        return
            else:
                celda.value = value
        except Exception:
            try:
                ws.cell(row=row, column=col).value = value
            except:
                pass
    
    # --- ORIGINAL: _marcar_checkbox (conservado) ---
    def _marcar_checkbox(self, ws, form_type: str):
        """Marca con 'X' el checkbox del tipo de formulario (método original)"""
        if not form_type:
            return
        
        # Buscar en las primeras filas donde están los checkboxes
        for r in range(1, 8):
            for c in range(1, 20):
                try:
                    cell = ws.cell(row=r, column=c)
                    
                    if isinstance(cell, MergedCell):
                        continue
                    
                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.upper().strip()
                        form_text = form_type.upper().strip()
                        
                        # Verificar coincidencia exacta
                        if form_text == cell_text or form_text in cell_text:
                            # Buscar celda de checkbox (puede estar a la izquierda o arriba)
                            for offset_c in range(-3, 2):
                                for offset_r in range(-1, 2):
                                    try:
                                        target_row = r + offset_r
                                        target_col = c + offset_c
                                        
                                        if target_col > 0 and target_row > 0:
                                            checkbox_cell = ws.cell(row=target_row, column=target_col)
                                            if not isinstance(checkbox_cell, MergedCell):
                                                val = str(checkbox_cell.value or "").strip()
                                                # Buscar celdas vacías o con símbolos de checkbox
                                                if val == "" or val in ["☐", "□", "[ ]", "[]"]:
                                                    self._escribir_celda_segura(ws, target_row, target_col, "X")
                                                    return
                                    except:
                                        continue
                except Exception:
                    continue

    # --- ADICIONAL: mapeo explícito para formularios y software (no destructivo) ---
    def _marcar_formularios_y_software(self, ws, data: UserData):
        """Marca tipos de formulario y casillas de software en celdas específicas."""
        # Mapeo explícito de texto de formulario a celda (columna letra, fila)
        form_cell_map = {
            "HOJA DE RESPONSABILIDAD": ("B", 6),
            "SOPORTE TÉCNICO": ("M", 6),
            "PASE DE SALIDA": ("V", 6),
            "DICTAMEN TÉCNICO": ("B", 7),
            "RECEPCION DE EQUIPO": ("M", 7),
            "RECEPCIÓN DE EQUIPO": ("M", 7),
            "PRÉSTAMO DE EQUIPO": ("V", 7),
            "PRESTAMO DE EQUIPO": ("V", 7),
            # Añade variantes si las necesitas
        }

        # Mapeo para sistemas y ofimática
        software_cell_map = {
            "WINDOWS": ("J", 38),
            "MAC": ("N", 38),
            "MS OFFICE": ("J", 39),
            "MSOFFICE": ("J", 39),
            "OFFICE": ("J", 39),
        }

        # Función auxiliar para convertir columna letra a índice
        def col_letter_to_index(letter: str) -> int:
            letter = letter.upper()
            idx = 0
            for ch in letter:
                idx = idx * 26 + (ord(ch) - ord('A') + 1)
            return idx

        # Limpiar previamente las celdas mapeadas (opcional)
        try:
            for _, (col_letter, row) in form_cell_map.items():
                self._escribir_celda_segura(ws, row, col_letter_to_index(col_letter), "")
            for _, (col_letter, row) in software_cell_map.items():
                self._escribir_celda_segura(ws, row, col_letter_to_index(col_letter), "")
        except Exception:
            pass

        # Marcar tipos de formulario (usa form_types del data)
        if getattr(data, "form_types", None):
            for ft in data.form_types:
                if not ft:
                    continue
                key = ft.strip().upper()
                # Normalizar acentos y espacios básicos
                key_norm = key.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
                if key_norm in form_cell_map:
                    col_letter, row = form_cell_map[key_norm]
                    col_idx = col_letter_to_index(col_letter)
                    try:
                        self._escribir_celda_segura(ws, row, col_idx, "X")
                    except Exception:
                        continue
                else:
                    # Intentar coincidencia parcial
                    for k_map in form_cell_map.keys():
                        if k_map in key_norm or key_norm in k_map:
                            col_letter, row = form_cell_map[k_map]
                            col_idx = col_letter_to_index(col_letter)
                            try:
                                self._escribir_celda_segura(ws, row, col_idx, "X")
                                break
                            except Exception:
                                continue

        # Marcar sistema operativo y ofimática según los campos del data
        so = (data.sistema_operativo or "").strip().upper()
        so_norm = so.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
        if so_norm:
            if "WINDOWS" in so_norm and "WINDOWS" in software_cell_map:
                col_letter, row = software_cell_map["WINDOWS"]
                self._escribir_celda_segura(ws, row, col_letter_to_index(col_letter), "X")
            if "MAC" in so_norm and "MAC" in software_cell_map:
                col_letter, row = software_cell_map["MAC"]
                self._escribir_celda_segura(ws, row, col_letter_to_index(col_letter), "X")

        ofi = (data.ofimática or "").strip().upper()
        ofi_norm = ofi.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
        if ofi_norm:
            if "MS OFFICE" in ofi_norm or "OFFICE" in ofi_norm or "MSOFFICE" in ofi_norm:
                if "MS OFFICE" in software_cell_map:
                    col_letter, row = software_cell_map["MS OFFICE"]
                    self._escribir_celda_segura(ws, row, col_letter_to_index(col_letter), "X")

    # --- ORIGINAL: _llenar_tabla_equipos (conservado) ---
    def _llenar_tabla_equipos(self, ws, equipos: list):
        """Llena la tabla de equipos adicionales (Sección C)"""
        if not equipos:
            return
        
        # Buscar la fila donde comienza la tabla de equipos
        tabla_inicio = None
        
        for row_idx in range(15, 45):
            for col_idx in range(1, 10):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value and isinstance(cell.value, str):
                        text = cell.value.upper()
                        if "EQUIPO" in text and "MARCA" in ws.cell(row=row_idx, column=col_idx+1).value.upper():
                            tabla_inicio = row_idx + 1
                            break
                except:
                    continue
            if tabla_inicio:
                break
        
        if not tabla_inicio:
            tabla_inicio = 25  # Posición por defecto
        
        # Llenar datos de equipos
        for i, equipo_info in enumerate(equipos):
            row_num = tabla_inicio + i
            try:
                # No, EQUIPO, MARCA, MODELO, SERIE, ESTADO
                self._escribir_celda_segura(ws, row_num, 1, str(i+1))
                self._escribir_celda_segura(ws, row_num, 2, equipo_info.get("equipo", ""))
                self._escribir_celda_segura(ws, row_num, 3, equipo_info.get("marca", ""))
                self._escribir_celda_segura(ws, row_num, 4, equipo_info.get("modelo", ""))
                self._escribir_celda_segura(ws, row_num, 5, equipo_info.get("serie", ""))
                self._escribir_celda_segura(ws, row_num, 6, equipo_info.get("estado", "OK"))
            except Exception:
                continue
    
    # --- ORIGINAL: _llenar_tabla_software (conservado) ---
    def _llenar_tabla_software(self, ws, software_list: list):
        """Llena la tabla de software adicional (Sección D)"""
        if not software_list:
            return
        
        # Buscar la tabla de software
        tabla_inicio = None
        
        for row_idx in range(25, 50):
            for col_idx in range(1, 10):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value and isinstance(cell.value, str):
                        text = cell.value.upper()
                        if "DESCRIPCIÓN" in text or "DESCRIPCION" in text:
                            tabla_inicio = row_idx + 1
                            break
                except:
                    continue
            if tabla_inicio:
                break
        
        if not tabla_inicio:
            return
        
        # Llenar software
        for i, software in enumerate(software_list):
            row_num = tabla_inicio + i
            try:
                # DESCRIPCIÓN, MARCA, VERSIÓN, OBSERVACIONES
                self._escribir_celda_segura(ws, row_num, 1, software.get("descripcion", ""))
                self._escribir_celda_segura(ws, row_num, 2, software.get("marca", ""))
                self._escribir_celda_segura(ws, row_num, 3, software.get("version", ""))
                self._escribir_celda_segura(ws, row_num, 4, software.get("observaciones", ""))
            except Exception:
                continue

# --- ADICIONAL / STUB: integración con Teams (esqueleto) ---
def send_file_to_teams_stub(local_path: str, teams_target: Dict[str, str]) -> bool:
    """
    STUB: Enviar archivo a Teams (no implementado).
    teams_target puede contener:
      - mode: "user" | "channel" | "chat"
      - target_id: id del canal/team o correo del usuario
      - message: texto opcional
    Instrucciones para implementar:
      1) Obtener token con MSAL (client credentials o delegated).
      2) Subir archivo a OneDrive o al drive del canal (SharePoint).
      3) Enviar mensaje en canal/chat referenciando el driveItem.
      4) Manejar permisos y errores.
    """
    messagebox.showinfo("Enviar a Teams (stub)", f"Se intentaría enviar:\n{os.path.basename(local_path)}\na {teams_target}")
    return True

# --- ORIGINAL: FormularioApp (conservado y extendido no destructivamente) ---
class FormularioApp:
    """Aplicación principal con interfaz gráfica"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Formulario DISAGRO - Departamento de Sistemas")
        self.root.geometry("900x700")
        self.data = UserData()
        self.excel_gen = ExcelGenerator()
        self.widgets = {}
        
        self._crear_interfaz()
    
    def _crear_interfaz(self):
        """Crea todos los elementos de la interfaz"""
        
        title_frame = tk.Frame(self.root, bg="#2c3e50", padx=10, pady=10)
        title_frame.grid(row=0, column=0, sticky="ew")
        
        tk.Label(
            title_frame,
            text="SISTEMA DE FORMULARIOS - DEPARTAMENTO DE SISTEMAS",
            font=("Arial", 14, "bold"),
            bg="#2c3e50",
            fg="white"
        ).pack()
        
        # Frame de búsqueda
        search_frame = tk.LabelFrame(self.root, text="1. Buscar Usuario", padx=10, pady=10)
        search_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
        
        tk.Label(search_frame, text="Alias o correo:").grid(row=0, column=0, sticky="w")
        self.widgets['alias_entry'] = tk.Entry(search_frame, width=40)
        self.widgets['alias_entry'].grid(row=0, column=1, padx=5)
        
        tk.Button(
            search_frame,
            text="🔍 Buscar en Outlook",
            command=self._buscar_usuario,
            bg="#3498db",
            fg="white",
            padx=10
        ).grid(row=0, column=2)
        
        self._crear_treeview()
        
        button_frame = tk.Frame(self.root, padx=10, pady=10)
        button_frame.grid(row=3, column=0)
        
        self.widgets['btn_rellenar'] = tk.Button(
            button_frame,
            text="📝 Rellenar Datos del Formulario",
            command=self._abrir_formulario_datos,
            state="disabled",
            bg="#e67e22",
            fg="white",
            padx=20,
            pady=10,
            font=("Arial", 10, "bold")
        )
        self.widgets['btn_rellenar'].grid(row=0, column=0, padx=5)
        
        tk.Button(
            button_frame,
            text="✅ Generar Plantilla Excel",
            command=self._generar_plantilla,
            bg="#27ae60",
            fg="white",
            padx=20,
            pady=10,
            font=("Arial", 10, "bold")
        ).grid(row=0, column=1, padx=5)
    
    def _crear_treeview(self):
        """Crea el TreeView para mostrar resultados"""
        tree_frame = tk.LabelFrame(self.root, text="2. Usuario Encontrado", padx=10, pady=5)
        tree_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=5)
        
        columns = ("Nombre", "Correo", "Puesto", "Empresa", "Departamento")
        self.widgets['tree'] = ttk.Treeview(
            tree_frame,
            columns=columns,
            show="headings",
            height=3
        )
        
        for col in columns:
            self.widgets['tree'].heading(col, text=col)
            self.widgets['tree'].column(col, width=160)
        
        self.widgets['tree'].pack(fill="both", expand=True)
    
    # --- ORIGINAL: _buscar_usuario (conservado) ---
    def _buscar_usuario(self):
        """Busca usuario en Outlook"""
        alias = self.widgets['alias_entry'].get().strip()
        
        if not alias:
            messagebox.showwarning("Campo vacío", "Ingresa un alias o correo.")
            return
        
        tree = self.widgets['tree']
        tree.delete(*tree.get_children())
        
        resultado = OutlookSearcher.buscar_usuario(alias)
        
        if resultado:
            for key, value in resultado.items():
                setattr(self.data, key, value)
            
            tree.insert("", "end", values=(
                resultado["full_name"],
                resultado["correo"],
                resultado["puesto"],
                resultado["empresa"],
                resultado["depart"]
            ))
            
            self.widgets['btn_rellenar'].config(state="normal")
            messagebox.showinfo("✓ Usuario encontrado", f"Usuario {resultado['full_name']} cargado.")
        else:
            messagebox.showwarning(
                "No encontrado",
                f"No se encontró el usuario: {alias}"
            )
    
    # --- ORIGINAL: _abrir_formulario_datos (conservado y extendido con opciones no destructivas) ---
    def _abrir_formulario_datos(self):
        """Abre ventana para capturar datos adicionales"""
        
        if not self.data.alias:
            messagebox.showwarning("Busca primero", "Primero debes buscar un usuario.")
            return
        
        win = tk.Toplevel(self.root)
        win.title(f"Datos del Formulario - {self.data.full_name}")
        win.geometry("800x900")
        
        canvas = tk.Canvas(win)
        scrollbar = tk.Scrollbar(win, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        entries = {}
        row = 0
        
        # ===== ENCABEZADO =====
        tk.Label(scrollable_frame, text="DATOS DEL FORMULARIO", font=("Arial", 12, "bold"), 
                 bg="#2c3e50", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        row += 1
        
        # Ticket y fecha
        tk.Label(scrollable_frame, text="No. de Ticket *").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['ticket'] = tk.Entry(scrollable_frame, width=50)
        entries['ticket'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['ticket'].insert(0, self.data.ticket)
        row += 1
        
        tk.Label(scrollable_frame, text="Fecha").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['fecha'] = tk.Entry(scrollable_frame, width=50)
        entries['fecha'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['fecha'].insert(0, datetime.now().strftime("%d/%m/%Y"))
        row += 1
        
        # ===== TIPO DE FORMULARIO =====
        tk.Label(scrollable_frame, text="TIPO DE FORMULARIO", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        tk.Label(scrollable_frame, text="Selecciona tipo(s) *").grid(row=row, column=0, sticky="ne", padx=5, pady=5)
        
        form_frame = tk.Frame(scrollable_frame)
        form_frame.grid(row=row, column=1, sticky="w", padx=5, pady=5)
        
        form_vars = {}
        for i, form_type in enumerate(ExcelGenerator.FORM_TYPES):
            var = tk.BooleanVar()
            if form_type in self.data.form_types:
                var.set(True)
            cb = tk.Checkbutton(form_frame, text=form_type, variable=var)
            cb.grid(row=i, column=0, sticky="w")
            form_vars[form_type] = var
        
        entries['form_types'] = form_vars
        row += 1
        
        # ===== SECCIÓN A: DATOS DEL COLABORADOR =====
        tk.Label(scrollable_frame, text="A. DATOS DEL COLABORADOR", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        tk.Label(scrollable_frame, text="(Ya cargados desde Outlook)", font=("Arial", 9, "italic"), 
                 fg="gray").grid(row=row, column=0, columnspan=2, pady=2)
        row += 1
        
        # Campos adicionales de Sección A
        campos_a = [
            ("9. Localidad", "localidad"),
            ("10. País", "pais"),
        ]
        
        for label, key in campos_a:
            tk.Label(scrollable_frame, text=label).grid(row=row, column=0, sticky="e", padx=5, pady=5)
            entries[key] = tk.Entry(scrollable_frame, width=50)
            entries[key].grid(row=row, column=1, sticky="w", padx=5, pady=5)
            entries[key].insert(0, getattr(self.data, key, ""))
            row += 1
        
        # ===== SECCIÓN B: HARDWARE =====
        tk.Label(scrollable_frame, text="B. HARDWARE (Campos 11-19)", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        campos_hardware = [
            ("11. Tipo de Computadora *", "pc_type"),
            ("12. Marca *", "marca"),
            ("13. Modelo", "model"),
            ("14. Service Tag (ST) *", "st"),
            ("15. Fecha de Compra", "fecha_compra"),
            ("16. Disco Duro", "disco_duro"),
            ("17. Memoria RAM", "memoria_ram"),
            ("18. IP", "ip"),
            ("19. Host Name", "hostname"),
        ]
        
        for label, key in campos_hardware:
            tk.Label(scrollable_frame, text=label).grid(row=row, column=0, sticky="e", padx=5, pady=5)
            entries[key] = tk.Entry(scrollable_frame, width=50)
            entries[key].grid(row=row, column=1, sticky="w", padx=5, pady=5)
            entries[key].insert(0, getattr(self.data, key, ""))
            row += 1
        
        # Estado del equipo
        tk.Label(scrollable_frame, text="Estado del equipo").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['estado'] = ttk.Combobox(
            scrollable_frame,
            values=["OK", "MALO", "REGULAR"],
            state="readonly",
            width=47
        )
        entries['estado'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['estado'].set(self.data.estado)
        row += 1
        
        # ===== SECCIÓN C: EQUIPOS ADICIONALES =====
        tk.Label(scrollable_frame, text="C. EQUIPOS ADICIONALES", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        equipos_frame = tk.Frame(scrollable_frame, relief="sunken", bd=1)
        equipos_frame.grid(row=row, column=0, columnspan=2, padx=5, pady=5, sticky="ew")
        
        equipos_list = tk.Listbox(equipos_frame, height=5, width=80)
        equipos_list.pack(side="left", fill="both", expand=True)
        
        equipos_scroll = tk.Scrollbar(equipos_frame, command=equipos_list.yview)
        equipos_scroll.pack(side="right", fill="y")
        equipos_list.config(yscrollcommand=equipos_scroll.set)
        
        if self.data.more_equipment:
            for eq in self.data.more_equipment:
                equipos_list.insert("end", f"{eq.get('equipo', '')} - {eq.get('marca', '')} {eq.get('modelo', '')} - {eq.get('serie', '')}")
        
        btn_equipos_frame = tk.Frame(scrollable_frame)
        btn_equipos_frame.grid(row=row+1, column=0, columnspan=2)
        
        def agregar_equipo():
            eq_win = tk.Toplevel(win)
            eq_win.title("Agregar Equipo Adicional")
            eq_win.geometry("450x300")
            
            eq_entries = {}
            labels = ["Equipo:", "Marca:", "Modelo:", "Serie:", "Estado:"]
            keys = ["equipo", "marca", "modelo", "serie", "estado"]
            
            for i, (lbl, key) in enumerate(zip(labels, keys)):
                tk.Label(eq_win, text=lbl).grid(row=i, column=0, sticky="e", padx=5, pady=5)
                if key == "estado":
                    e = ttk.Combobox(eq_win, values=["OK", "MALO", "REGULAR"], state="readonly", width=28)
                    e.set("OK")
                else:
                    e = tk.Entry(eq_win, width=30)
                e.grid(row=i, column=1, padx=5, pady=5)
                eq_entries[key] = e
            
            def guardar_equipo():
                equipo_data = {}
                for k, v in eq_entries.items():
                    equipo_data[k] = v.get().strip()
                
                if equipo_data["equipo"]:
                    equipos_list.insert("end", f"{equipo_data['equipo']} - {equipo_data['marca']} {equipo_data['modelo']} - {equipo_data['serie']}")
                    if not hasattr(self.data, 'temp_equipos'):
                        self.data.temp_equipos = []
                    self.data.temp_equipos.append(equipo_data)
                    eq_win.destroy()
                else:
                    messagebox.showwarning("Campo vacío", "El nombre del equipo es obligatorio")
            
            tk.Button(eq_win, text="💾 Guardar", command=guardar_equipo, bg="#27ae60", fg="white").grid(row=len(labels), column=0, columnspan=2, pady=10)
        
        def eliminar_equipo():
            sel = equipos_list.curselection()
            if not sel:
                messagebox.showwarning("Selecciona", "Selecciona un equipo para eliminar")
                return
            idx = sel[0]
            equipos_list.delete(idx)
            if hasattr(self.data, 'temp_equipos') and idx < len(self.data.temp_equipos):
                try:
                    self.data.temp_equipos.pop(idx)
                except:
                    pass
        
        tk.Button(btn_equipos_frame, text="➕ Agregar", command=agregar_equipo, bg="#3498db", fg="white").grid(row=0, column=0, padx=5, pady=5)
        tk.Button(btn_equipos_frame, text="➖ Eliminar", command=eliminar_equipo, bg="#c0392b", fg="white").grid(row=0, column=1, padx=5, pady=5)
        
        row += 3
        
        # ===== SECCIÓN D: SOFTWARE =====
        tk.Label(scrollable_frame, text="D. SOFTWARE", font=("Arial", 11, "bold"), 
                 bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15, 5))
        row += 1
        
        tk.Label(scrollable_frame, text="Sistema operativo").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['sistema_operativo'] = ttk.Combobox(scrollable_frame, values=["Windows", "Mac OS", ""], width=47)
        entries['sistema_operativo'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['sistema_operativo'].set(self.data.sistema_operativo)
        row += 1
        
        tk.Label(scrollable_frame, text="Ofimática").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['ofimatica'] = ttk.Combobox(scrollable_frame, values=["MS Office", "Open Office", ""], width=47)
        entries['ofimatica'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['ofimatica'].set(self.data.ofimática)
        row += 1
        
        # ===== ADICIONAL: Opciones de entrega (no destructivas) =====
        tk.Label(scrollable_frame, text="Opciones de entrega", font=("Arial", 11, "bold"), bg="#34495e", fg="white").grid(row=row, column=0, columnspan=2, sticky="ew", pady=(15,5))
        row += 1
        send_local_var = tk.BooleanVar(value=True)
        send_teams_var = tk.BooleanVar(value=False)
        tk.Checkbutton(scrollable_frame, text="Guardar copia local", variable=send_local_var).grid(row=row, column=0, sticky="w", padx=5, pady=5)
        tk.Checkbutton(scrollable_frame, text="Enviar copia a Teams", variable=send_teams_var).grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['send_local_var'] = send_local_var
        entries['send_teams_var'] = send_teams_var
        row += 1
        
        tk.Label(scrollable_frame, text="Ruta local (si aplica)").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['local_path'] = tk.Entry(scrollable_frame, width=50)
        entries['local_path'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['local_path'].insert(0, os.getcwd())
        tk.Button(scrollable_frame, text="📁", command=lambda: entries['local_path'].delete(0, 'end') or entries['local_path'].insert(0, filedialog.askdirectory())).grid(row=row, column=2, sticky="w")
        row += 1
        
        tk.Label(scrollable_frame, text="Destino Teams (correo o channel id)").grid(row=row, column=0, sticky="e", padx=5, pady=5)
        entries['teams_target'] = tk.Entry(scrollable_frame, width=50)
        entries['teams_target'].grid(row=row, column=1, sticky="w", padx=5, pady=5)
        entries['teams_target'].insert(0, "")
        tk.Button(scrollable_frame, text="Probar envío (stub)", command=lambda: send_file_to_teams_stub("prueba.xlsx", {"mode":"user","target_id":entries['teams_target'].get(),"message":"Prueba desde app"})).grid(row=row, column=2, sticky="w", padx=5)
        row += 1
        
        # Botones de guardar formulario
        def guardar_formulario():
            # Validaciones mínimas
            ticket_val = entries['ticket'].get().strip()
            if not ticket_val:
                messagebox.showwarning("Faltan datos", "El número de ticket es obligatorio")
                return
            # Guardar campos en self.data
            self.data.ticket = ticket_val
            self.data.fecha = entries['fecha'].get().strip()
            # form types
            selected_forms = []
            for k, var in entries['form_types'].items():
                if var.get():
                    selected_forms.append(k)
            self.data.form_types = selected_forms
            # Sección A
            self.data.localidad = entries['localidad'].get().strip()
            self.data.pais = entries['pais'].get().strip()
            # Hardware
            for _, key in campos_hardware:
                setattr(self.data, key, entries[key].get().strip())
            self.data.estado = entries['estado'].get().strip()
            # Software
            self.data.sistema_operativo = entries['sistema_operativo'].get().strip()
            self.data.ofimática = entries['ofimatica'].get().strip()
            # Equipos temporales ya guardados en self.data.temp_equipos
            messagebox.showinfo("Guardado", "Datos del formulario guardados en memoria.")
            win.destroy()
        
        tk.Button(scrollable_frame, text="💾 Guardar formulario", command=guardar_formulario, bg="#27ae60", fg="white", padx=10, pady=8).grid(row=row, column=0, columnspan=2, pady=15)
    
    # --- ORIGINAL: _generar_plantilla (conservado y adaptado para opciones) ---
    def _generar_plantilla(self):
        # Antes de generar, asegurarse de que ticket y alias estén presentes
        if not self.data.alias:
            messagebox.showwarning("Busca primero", "Primero debes buscar un usuario.")
            return
        if not self.data.ticket:
            messagebox.showwarning("Falta ticket", "Debes completar el número de ticket.")
            return
        
        # Preguntar carpeta de guardado local
        out_dir = filedialog.askdirectory(title="Selecciona carpeta para guardar copia local", initialdir=os.getcwd())
        if not out_dir:
            if not messagebox.askyesno("Continuar", "No seleccionaste carpeta. ¿Deseas guardar en la carpeta actual?"):
                return
            out_dir = os.getcwd()
        
        output_file = self.excel_gen.generar(self.data, output_dir=out_dir)
        if not output_file:
            return
        
        # Si el usuario quiere enviar a Teams, pedir destino y usar stub
        if messagebox.askyesno("Enviar a Teams", "¿Deseas enviar una copia a Teams ahora?"):
            dest = simpledialog.askstring("Destino Teams", "Ingresa correo del usuario o channel id:")
            if dest:
                teams_target = {"mode": "user", "target_id": dest, "message": f"Plantilla {os.path.basename(output_file)}"}
                try:
                    send_file_to_teams_stub(output_file, teams_target)
                except Exception as e:
                    messagebox.showerror("Error envío Teams", f"No se pudo enviar a Teams:\n{str(e)}")
        
        messagebox.showinfo("Guardado exitoso", f"Plantilla exportada como:\n{output_file}")
        # limpiar temp_equipos si quieres
        self.data.temp_equipos = []

# --- MAIN ---
if __name__ == "__main__":
    root = tk.Tk()
    app = FormularioApp(root)
    root.mainloop()
